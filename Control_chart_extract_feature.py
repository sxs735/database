#%%
import re, csv
from pathlib import Path
from typing import Any, Dict, List, Tuple
from datetime import datetime
import numpy as np
from openpyxl import Workbook, load_workbook
from scipy.optimize import curve_fit


SUPPORTED_EXTENSIONS = {".csv", ".txt", ".s2p"}
MAIN_PATTERN = re.compile(r"""
                          ^(?P<datatype>[^_]+)
                          _(?P<wafer>[^_]+)
                          _(?P<doe>[^_]+)
                          _(?P<cage>[^_]+)
                          _die(?P<die>\d+)
                          _(?P<temperature>-?\d+)C
                          _\#(?P<repeat>\d+)
                          _(?P<device>[^_]+)
                          _ch_(?P<ch_in>\d+)
                          _(?P<ch_out>\d+)
                          _(?P<power>-?\d+)dBm
                          (?P<rest>.*)
                          \.(?:csv|txt|s2p)$""",
                          re.VERBOSE)

def tofloat(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return np.nan

def parse_filename(filename: str) -> Dict[str, Any]:
    name = Path(filename).name
    match = MAIN_PATTERN.match(name)

    if not match:
        raise ValueError(f"檔名格式不符: {filename}")

    result = match.groupdict()
    rest = result.pop("rest")

    result["SMU"] = []
    result["arguments"] = []

    if not rest:
        return result

    tokens = rest.strip("_").split("_")
    i = 0
    pass_SMU = False
    while i < len(tokens):
        token = tokens[i]
        if token == "SMU":
            match = re.match(r"([-+]?\d*\.?\d+)([a-zA-Z%]*)", tokens[i + 3])
            #print(match.groups())
            result["SMU"].append({"element": tokens[i + 1],
                                    "channel": tokens[i + 2],
                                    "set_mode": 'VOLT' if match[2] in ['V', 'mV'] else 'CURR',
                                    "set_value": tokens[i + 3]})
            i += 4
            continue
        if i + 2 < len(tokens) and token != "arg" and not pass_SMU:
            match = re.match(r"([-+]?\d*\.?\d+)([a-zA-Z%]*)", tokens[i + 2])
            result["SMU"].append({"element": tokens[i],
                                    "channel": tokens[i + 1],
                                    "set_mode": 'VOLT' if match[2] in ['V', 'mV'] else 'CURR',
                                    "set_value": tokens[i + 2]})
            i += 3
            continue
        if token == "arg":
            result["arguments"].append({f"arg": tokens[i + 1]})
            i += 2
            pass_SMU = True
            continue
        if token:
            result["arguments"].append({"arg": token})
        i += 1

    return result

def parse_folder(folder_path: str) -> Tuple[Dict[Path, Dict[str, Any]], List[str], float, float]:
    """
    檢測資料夾內所有檔案名稱是否符合指定格式

    Returns:
    - (valid_files, invalid_files) 元組
        - valid_files: {Path物件: 解析後的元數據字典, ...}
        - invalid_files: 不符合格式的檔案名稱列表
    """
    folder = Path(folder_path)

    if not folder.exists():
        raise FileNotFoundError(f"資料夾不存在: {folder_path}")

    valid_files: Dict[Path, Dict[str, Any]] = {}
    invalid_files: List[str] = []

    birthtime = []
    for ext in SUPPORTED_EXTENSIONS:
        for file_path in folder.glob(f"*{ext}"):
            try:
                birthtime.append(file_path.stat().st_mtime)
                meta = parse_filename(file_path.name)
                valid_files[file_path] = meta
            except ValueError:
                invalid_files.append(file_path.name)
    if len(valid_files) == 0:
        print("資料夾內沒有符合格式的檔案。")
        print('testing filename parsing')
        try:
            file_path = list(folder.glob("*.csv"))[0]
        except IndexError:
            file_path = list(folder.glob("*.s2p"))[0]
        test_filename_parsing(file_path.name)
        raise ValueError("資料夾內沒有符合格式的檔案。")

    return valid_files, invalid_files, min(birthtime), max(birthtime)

def test_filename_parsing(filename: str) -> None:
    patterns = [("datatype", r"^[^_]+"),
                ("wafer", r"_[^_]+"),
                ("doe", r"_[^_]+"),
                ("cage", r"_[^_]+"),
                ("die", r"_die\d+"),
                ("temperature", r"_-?\d+C"),
                ("repeat", r"_\#\d+"),
                ("device", r"_[^_]+"),
                ("channel", r"_ch_\d+_\d+"),
                ("power", r"_-?\d+dBm"),
                ("rest", r".*\.(?:csv|txt|s2p)$")]

    pos = 0
    for name, pattern in patterns:
        m = re.match(pattern, filename[pos:])
        if not m:
            print(f"Mismatch at {name}")
            break
        else:
            print(f"{name}: pass")
        pos += m.end()

def read_spectrum_all(path):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        mode = None
        header = []
        Min_Max = []
        Mueller = []
        Avg = []
        PDL = []
        TE_TM = []
        for i,row in enumerate(reader):
            #print(row)
            if '=== Min' in row and mode != 'min_max':
                mode = 'min_max'
            elif '=== Average IL (TLS 0) ===' in row and mode != 'average_il':
                mode = 'average_il'
            elif '=== Mueller Row 1 (TLS 0) ===' in row and mode != 'mueller':
                mode = 'mueller'
            elif '=== PDL (TLS 0) ===' in row and mode != 'pdl':
                mode = 'pdl'
            elif '=== TE' in row and mode != 'te_tm':
                mode = 'te_tm'
            elif mode == 'min_max':
                Min_Max += [[tofloat(value) for value in row]]
            elif mode == 'average_il':
                Avg += [[tofloat(value) for value in row]]
            elif mode == 'mueller':
                Mueller += [[tofloat(value) for value in row]]
            elif mode == 'pdl':
                PDL += [[tofloat(value) for value in row]]
            elif mode == 'te_tm':
                TE_TM += [[tofloat(value) for value in row]]
            else:
                header += [row]
        data = {'min_max': np.array(Min_Max),
                'average_il': np.array(Avg),
                'mueller': np.array(Mueller),
                'pdl': np.array(PDL),
                'te_tm': np.array(TE_TM),
                'header': header}
    return data

def extract_feature(spcm_data,spcm_ref,deg=5,range_nm=(1280,1340)):
    #Output: [[peak_wavelength, peak_loss, bandwidth], [peak_wavelength, peak_loss, bandwidth]]
    x_array = np.arange(range_nm[0], range_nm[1], 0.001)
    x_norm = (x_array - range_nm[0]) / (range_nm[1] - range_nm[0])

    spcm_data[:,0] = spcm_data[:,0]*1E9
    spcm_ref[:,0] = spcm_ref[:,0]*1E9
    coef1 = np.polyfit((spcm_data[:,0]-range_nm[0])/(range_nm[1]-range_nm[0]), -(spcm_data[:,1]-spcm_ref[:,1]), deg=deg)
    coef2 = np.polyfit((spcm_data[:,0]-range_nm[0])/(range_nm[1]-range_nm[0]), -(spcm_data[:,3]-spcm_ref[:,3]), deg=deg)
    spcm1 = np.polyval(coef1, x_norm) 
    spcm2 = np.polyval(coef2, x_norm)
    peak_idx_1 = np.argmax(spcm1)
    peak_idx_2 = np.argmax(spcm2)
    x0_1, y0_1 = float(x_array[peak_idx_1]), float(spcm1[peak_idx_1])
    x0_2, y0_2 = float(x_array[peak_idx_2]), float(spcm2[peak_idx_2])
    left_idx_1 = np.abs(spcm1[:peak_idx_1] - (y0_1 - 1)).argmin()
    right_idx_1 = np.abs(spcm1[peak_idx_1:] - (y0_1 - 1)).argmin() + peak_idx_1
    left_idx_2 = np.abs(spcm2[:peak_idx_2] - (y0_2 - 1)).argmin()
    right_idx_2 = np.abs(spcm2[peak_idx_2:] - (y0_2 - 1)).argmin() + peak_idx_2
    x1_1,x2_1 = float(x_array[left_idx_1]), float(x_array[right_idx_1])
    x1_2,x2_2 = float(x_array[left_idx_2]), float(x_array[right_idx_2])
    return [[x0_1,y0_1, x2_1-x1_1], [x0_2, y0_2, x2_2-x1_2]]

def write_metric_row(xlsx_path: Path, sheet_name: str, day: str, values, header_row) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(xlsx_path) if xlsx_path.exists() else Workbook()
    except Exception:
        wb = Workbook()

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    # 1) 依據外部給予 header_row（可不含 day）
    incoming_headers = [str(h) for h in header_row if h is not None]
    if len(incoming_headers) == 0:
        raise ValueError("header_row 不可為空")
    if incoming_headers[0].lower() != "day":
        incoming_headers = ["day"] + incoming_headers

    # 先讀取既有 header
    existing_headers = []
    if ws.max_row >= 1:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is None:
                continue
            existing_headers.append(str(v))

    # 初始化空白表頭
    if len(existing_headers) == 0:
        existing_headers = ["day"]

    # 確保 day 欄位存在且在第 1 欄
    if "day" not in [h.lower() for h in existing_headers]:
        existing_headers = ["day"] + existing_headers
    elif existing_headers[0].lower() != "day":
        day_idx = [h.lower() for h in existing_headers].index("day")
        existing_headers.insert(0, existing_headers.pop(day_idx))

    # 2) header 已存在 -> 覆蓋對應 values
    # 3) header 不存在 -> 橫向新增 header
    incoming_data_headers = incoming_headers[1:]
    for h in incoming_data_headers:
        if h not in existing_headers:
            existing_headers.append(h)

    # 回寫完整 header 列
    for col, value in enumerate(existing_headers, start=1):
        ws.cell(row=1, column=col, value=value)

    arr = np.asarray(values, dtype=float).tolist()
    if len(arr) != len(incoming_data_headers):
        raise ValueError(f"values 長度({len(arr)})需等於 header_row(扣除 day) 長度({len(incoming_data_headers)})")

    value_map = {h: v for h, v in zip(incoming_data_headers, arr)}

    target_row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == str(day):
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1

    # 寫 day
    ws.cell(row=target_row, column=1, value=day)

    # 依 header 對應欄位寫入/覆蓋 values
    for c, h in enumerate(existing_headers, start=1):
        if c == 1:
            continue
        if h in value_map:
            ws.cell(row=target_row, column=c, value=value_map[h])

    wb.save(xlsx_path)
    wb.close()

def asymmetric_super_gaussian(x, amplitude, center, width_left, width_right, order, offset):
    """
    Asymmetric Super-Gaussian function.

    Parameters:
        x: wavelength array
        amplitude: peak amplitude (negative for insertion loss minimum)
        center: peak center wavelength
        width_left: width parameter for left side (x <= center)
        width_right: width parameter for right side (x > center)
        order: super-gaussian order (higher = flatter top)
        offset: baseline offset

    Returns:
        y values at x positions
    """
    result = np.zeros_like(x, dtype=float)
    left_mask = x <= center
    right_mask = x > center
    result[left_mask] = amplitude * np.exp(-(np.abs((x[left_mask] - center) / width_left) ** order)) + offset
    result[right_mask] = amplitude * np.exp(-(np.abs((x[right_mask] - center) / width_right) ** order)) + offset
    return result

def fit_insertion_loss_spectrum(wavelength, optical_power, fitted_wavelength, fitting_window=30):
    """
    Standard curve fitting function for optical insertion loss spectra.

    Uses Asymmetric Super-Gaussian (validated method).

    Parameters:
        wavelength (array): Wavelength array in nanometers (nm)
        optical_power (array): Insertion loss in dB (lower = better transmission)
        fitted_wavelength (list): Wavelength at which to calculate fitted power (nm)
        fitting_window (float, optional): Fitting window size in nm. Default: 30
            - Positive value: window size in nm (e.g., 30 for ±15nm window)
            - -1: fit all data points (no window)
            - 0 or negative (except -1): same as -1

    Returns:
        fitted_optical_power (float): Fitted insertion loss at fitted_wavelength (dB)
        extracted_parameters (dict): Dictionary containing:
            - 'fitted_peak_wavelength' (float): Wavelength at minimum insertion loss (nm)
            - 'fitted_peak_power' (float): Minimum insertion loss value (dB)
            - 'BW' (float): 2dB bandwidth (nm)
            - 'BW_start_wavelength' (float): Left edge of 2dB bandwidth (nm)
            - 'BW_end_wavelength' (float): Right edge of 2dB bandwidth (nm)
            - 'fit_success' (bool): Whether fitting succeeded
            - 'fit_params' (dict): Fitted parameters (amplitude, center, etc.)
            - 'fitting_window_used' (float): Actual window size used (nm), -1 if full range

    Examples:
        >>> # Standard 30nm window (default, validated)
        >>> fitted_power, params = fit_insertion_loss_spectrum(wavelength, power, 1308.0)

        >>> # Custom 40nm window
        >>> fitted_power, params = fit_insertion_loss_spectrum(wavelength, power, 1308.0, fitting_window=40)

        >>> # Fit all data points (no window)
        >>> fitted_power, params = fit_insertion_loss_spectrum(wavelength, power, 1308.0, fitting_window=-1)
    """

    # Initialize output
    extracted_parameters = {
        'fitted_peak_wavelength': np.nan,
        'fitted_peak_power': np.nan,
        'BW': np.nan,
        'BW_start_wavelength': np.nan,
        'BW_end_wavelength': np.nan,
        'fit_success': False,
        'fit_params': {},
        'fitting_window_used': fitting_window if fitting_window > 0 else -1
    }

    if optical_power is None:
        return np.nan, extracted_parameters
    # Clean data (remove NaN/inf)
    mask = np.isfinite(wavelength) & np.isfinite(optical_power)
    x_clean = wavelength[mask]
    y_clean = optical_power[mask]

    if len(x_clean) < 10:
        print("Warning: Insufficient valid data points (< 10)")
        return np.nan, extracted_parameters

    # Step 1: Find raw peak (minimum insertion loss)
    approx_peak_idx = np.argmin(y_clean)
    approx_peak = x_clean[approx_peak_idx]
    approx_peak_value = y_clean[approx_peak_idx]

    # Step 2: Determine fitting window
    if fitting_window > 0:
        # Use specified window size
        window_half_width = fitting_window / 2.0
        window_left = approx_peak - window_half_width
        window_right = approx_peak + window_half_width

        # Step 3: Extract data within window
        range_mask = (x_clean >= window_left) & (x_clean <= window_right)
        x_fit = x_clean[range_mask]
        y_fit = y_clean[range_mask]

        if len(x_fit) < 10:
            print(f"Warning: Insufficient data points in {fitting_window}nm window (< 10)")
            return np.nan, extracted_parameters
    else:
        # Fit all data points (no window)
        x_fit = x_clean
        y_fit = y_clean
        window_left = x_clean.min()
        window_right = x_clean.max()

    # Step 4: Fit Asymmetric Super-Gaussian to windowed data
    try:
        # Initial guess
        amplitude_guess = np.min(y_fit) - np.max(y_fit)  # Negative for minimum
        center_guess = approx_peak
        width_guess = 5.0  # nm
        order_guess = 2.0
        offset_guess = np.max(y_fit)

        p0 = [amplitude_guess, center_guess, width_guess, width_guess, order_guess, offset_guess]

        # Parameter bounds
        bounds = (
            [-np.inf, window_left, 0.1, 0.1, 1.0, -np.inf],  # Lower bounds
            [0, window_right, 20.0, 20.0, 10.0, np.inf]      # Upper bounds
        )

        # Perform fitting
        popt, pcov = curve_fit(
            asymmetric_super_gaussian,
            x_fit,
            y_fit,
            p0=p0,
            bounds=bounds,
            maxfev=10000
        )

        amplitude, center, width_left_fit, width_right_fit, order, offset = popt

        # Store fit parameters
        extracted_parameters['fit_params'] = {
            'amplitude': amplitude,
            'center': center,
            'width_left': width_left_fit,
            'width_right': width_right_fit,
            'order': order,
            'offset': offset
        }

        # Create fitted function
        def fit_func(x):
            return asymmetric_super_gaussian(x, *popt)

        extracted_parameters['fit_success'] = True

    except Exception as e:
        print(f"Warning: Fitting failed: {e}")
        return np.nan, extracted_parameters

    # Step 5: Extract fitted peak wavelength and power
    # Search for minimum in fitted function over full data range
    x_search = np.linspace(x_clean.min(), x_clean.max(), 2000)
    y_search = fit_func(x_search)
    peak_idx = np.argmin(y_search)

    fitted_peak_wavelength = x_search[peak_idx]
    fitted_peak_power = y_search[peak_idx]

    extracted_parameters['fitted_peak_wavelength'] = fitted_peak_wavelength
    extracted_parameters['fitted_peak_power'] = fitted_peak_power

    # Step 6: Calculate fitted optical power at input wavelength
    fitted_optical_power = fit_func(np.array(fitted_wavelength))

    # Step 7: Calculate 2dB bandwidth
    threshold = fitted_peak_power + 2.0  # 2dB above minimum

    # Evaluate fitted function over full range
    x_full = np.linspace(x_clean.min(), x_clean.max(), 2000)
    y_full = fit_func(x_full)

    # Find where curve is below threshold (good transmission region)
    below_threshold = y_full <= threshold

    if np.any(below_threshold):
        indices = np.where(below_threshold)[0]

        if len(indices) >= 2:
            # Bandwidth edges
            BW_start_wavelength = x_full[indices[0]]
            BW_end_wavelength = x_full[indices[-1]]
            BW = BW_end_wavelength - BW_start_wavelength

            extracted_parameters['BW'] = BW
            extracted_parameters['BW_start_wavelength'] = BW_start_wavelength
            extracted_parameters['BW_end_wavelength'] = BW_end_wavelength
        else:
            print("Warning: Could not determine 2dB bandwidth edges")
    else:
        print("Warning: No points below 2dB threshold")

    return fitted_optical_power, extracted_parameters

#%%
group_tf = {('1','1'): ['Group1','Group6'], 
            ('2','2'): ['Group2','Group7'], 
            ('3','3'): ['Group3','Group8'], 
            ('4','4'): ['Group4','Group9'],
            ('5','5'): ['Group5','Group10'],
            ('6','6'): ['Group11','Group16'],
            ('7','7'): ['Group12','Group17'],
            ('8','8'): ['Group13','Group18'],
            ('9','9'): ['Group14','Group19'],
            ('10','10'): ['Group15','Group20'],
            ('11', '11'): ['RL15', 'Group21'],
            ('12', '12'): ['R15', 'NA'],
            ('12', '13'): ['NA', 'R16']}

measure_folder = Path(__file__).resolve().parent
#measure_folder = Path(r"D:\Data\1_DataBase\processing\mCoupe_J-FA04134196")
reference_file = measure_folder / "Ref_smooth_data.csv"
xlsx_path = measure_folder / "control_chat_table.xlsx"

#子資料夾名稱
subfolders = [p for p in measure_folder.iterdir() if p.is_dir()]
subfolder_time_pairs = []
for subfolder in subfolders:
    all_files = [f for f in subfolder.rglob("*") if f.is_file()]
    if len(all_files) != 0:
        earliest_ts = min(f.stat().st_mtime for f in all_files)
        subfolder_time_pairs.append((subfolder.name, earliest_ts))

# 排除沒有檔案的資料夾
#subfolder_time_pairs = [p for p in subfolder_time_pairs if p[1] is not None]
subfolder_time_pairs.sort(key=lambda x: x[1])
subfolder_name_list = [name for name, _ in subfolder_time_pairs]
wb = load_workbook(xlsx_path, data_only=True, read_only=True) if xlsx_path.exists() else Workbook()
ws = wb[wb.sheetnames[0]]
exist_day_list = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0] is not None]
wb.close()
unextracted_day_list = [day for day in subfolder_name_list if day not in exist_day_list]

for day in unextracted_day_list:
    data_info = parse_folder(measure_folder/day)[0]
    group = {}
    ref = {}
    for filtpath in data_info:
        spcm_info = data_info[filtpath]
        spcm_info['file_path'] = filtpath
        opt_ch = (spcm_info['ch_in'],spcm_info['ch_out'])
        if opt_ch not in group:
            group[opt_ch] = [spcm_info]
        else:
            group[opt_ch].append(spcm_info)

    if not reference_file.exists():
        raise FileNotFoundError(f"reference_file 不存在: {reference_file}")

    with open(reference_file, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"reference_file 無表頭: {reference_file}")

        ref = {h: [] for h in reader.fieldnames}
        for row in reader:
            for h in reader.fieldnames:
                ref[h].append(tofloat(row.get(h)))

    peak_wavelength = {}
    peak_loss = {}
    bandwidth = {}
    loss_fitted_wavelength = {}
    average_loss = {}
    for opt_ch in list(group.keys())[:]:
        repeat = []
        sampling =  [int(spcm_info['repeat']) for spcm_info in group[opt_ch]]
        arg_idx = np.argsort(sampling)
        for spcm_info in group[opt_ch]:
            path = spcm_info['file_path']
            spcm_data = read_spectrum_all(path)['min_max']
            group_name_1 = group_tf[opt_ch][0]
            group_name_2 = group_tf[opt_ch][1]
            #[x0_1,y0_1, bw_1], [x0_2, y0_2, bw_2] = extract_feature(spcm_data, spcm_ref, deg=5)
            wavelength = spcm_data[:,0]*1E9
            optical_power1 = (spcm_data[:,1]-ref[group_name_1]) if group_name_1 not in ['NA'] else None
            optical_power2 = (spcm_data[:,3]-ref[group_name_2]) if group_name_2 not in ['NA'] else None
            fitted_wavelength = [1300,1308,1320]
            fitted_optical_power1, extracted_parameters1 = fit_insertion_loss_spectrum(wavelength, optical_power1, fitted_wavelength, fitting_window=30)
            fitted_optical_power2, extracted_parameters2 = fit_insertion_loss_spectrum(wavelength, optical_power2, fitted_wavelength, fitting_window=30)
            x0_1 = extracted_parameters1['fitted_peak_wavelength']
            y0_1 = extracted_parameters1['fitted_peak_power']
            bw_1 = extracted_parameters1['BW']
            yf1 = extracted_parameters1['fitted_peak_power']
            x0_2 = extracted_parameters2['fitted_peak_wavelength']
            y0_2 = extracted_parameters2['fitted_peak_power']
            bw_2 = extracted_parameters2['BW']
            yf2 = extracted_parameters2['fitted_peak_power']
            y_avg_1 = np.mean(optical_power1) if optical_power1 is not None else None
            y_avg_2 = np.mean(optical_power2) if optical_power2 is not None else None
            if group_name_1 not in peak_wavelength:
                peak_wavelength[group_name_1] = [x0_1]
                peak_loss[group_name_1] = [y0_1]
                bandwidth[group_name_1] = [bw_1]
                loss_fitted_wavelength[group_name_1] = [fitted_optical_power1]
                average_loss[group_name_1] = [y_avg_1]
            else:
                peak_wavelength[group_name_1].append(x0_1)
                peak_loss[group_name_1].append(y0_1)
                bandwidth[group_name_1].append(bw_1)
                loss_fitted_wavelength[group_name_1].append(fitted_optical_power1)
                average_loss[group_name_1].append(y_avg_1)

            if group_name_2 not in peak_wavelength:
                peak_wavelength[group_name_2] = [x0_2]
                peak_loss[group_name_2] = [y0_2]
                bandwidth[group_name_2] = [bw_2]
                loss_fitted_wavelength[group_name_2] = [fitted_optical_power2]
                average_loss[group_name_2] = [y_avg_2]
            else:
                peak_wavelength[group_name_2].append(x0_2)
                peak_loss[group_name_2].append(y0_2)
                bandwidth[group_name_2].append(bw_2)
                loss_fitted_wavelength[group_name_2].append(fitted_optical_power2)
                average_loss[group_name_2].append(y_avg_2)
        sorted_sampling = list(np.asarray(sampling)[arg_idx])

        for key in [group_name_1, group_name_2]:
            if key in ['Group'+str(i) for i in [3,4,5,6,7,8,16,17,18,19,20,21]]:
                peak_wavelength[key] = np.asarray(peak_wavelength[key])[arg_idx]
                peak_loss[key] = np.asarray(peak_loss[key])[arg_idx]
                bandwidth[key] = np.asarray(bandwidth[key])[arg_idx]
                
                
                metric_header_row = ["day"] + [f"{key}_#{int(r)}" for r in sorted_sampling]
                write_metric_row(xlsx_path, "Short LB - Peak Wavelength", day, peak_wavelength[key], metric_header_row)
                write_metric_row(xlsx_path, "Short LB - Peak Loss", day, peak_loss[key], metric_header_row)
                write_metric_row(xlsx_path, "Short LB - Bandwidth", day, bandwidth[key], metric_header_row)
                write_metric_row(xlsx_path, "Short LB - Loss at 1300", day, loss_fitted_wavelength[key][0], metric_header_row)
                write_metric_row(xlsx_path, "Short LB - Loss at 1308", day, loss_fitted_wavelength[key][1], metric_header_row)
                write_metric_row(xlsx_path, "Short LB - Loss at 1320", day, loss_fitted_wavelength[key][2], metric_header_row)
            elif key =='Group1':
                peak_wavelength[key] = np.asarray(peak_wavelength[key])[arg_idx]
                peak_loss[key] = np.asarray(peak_loss[key])[arg_idx]
                bandwidth[key] = np.asarray(bandwidth[key])[arg_idx]
                
                metric_header_row = ["day"] + [f"{key}_#{int(r)}" for r in sorted_sampling]
                write_metric_row(xlsx_path, "G1 - Peak Wavelength", day, peak_wavelength[key], metric_header_row)
                write_metric_row(xlsx_path, "G1 - Peak Loss", day, peak_loss[key], metric_header_row)
                write_metric_row(xlsx_path, "G1 - Bandwidth", day, bandwidth[key], metric_header_row)
                write_metric_row(xlsx_path, "G1 - Loss at 1300", day, loss_fitted_wavelength[key][0], metric_header_row)
                write_metric_row(xlsx_path, "G1 - Loss at 1308", day, loss_fitted_wavelength[key][1], metric_header_row)
                write_metric_row(xlsx_path, "G1 - Loss at 1320", day, loss_fitted_wavelength[key][2], metric_header_row)
            elif key == 'R15':
                average_loss[key] = np.asarray(average_loss[key])[arg_idx]
                metric_header_row = ["day"] + [f"{key}_#{int(r)}" for r in sorted_sampling]
                write_metric_row(xlsx_path, "R15 - Average Loss", day, average_loss[key], metric_header_row)
            elif key == 'R16':
                average_loss[key] = np.asarray(average_loss[key])[arg_idx]
                metric_header_row = ["day"] + [f"{key}_#{int(r)}" for r in sorted_sampling]
                write_metric_row(xlsx_path, "R16 - Average Loss", day, average_loss[key], metric_header_row)
