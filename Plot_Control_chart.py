#%%
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib import font_manager
from matplotlib import rcParams
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime


candidates = ["Microsoft JhengHei","Microsoft YaHei","SimHei","Noto Sans CJK TC","PMingLiU","MingLiU",]
available_fonts = {f.name for f in font_manager.fontManager.ttflist}
selected = [name for name in candidates if name in available_fonts]
if selected:
    rcParams["font.sans-serif"] = selected + rcParams.get("font.sans-serif", [])
rcParams["axes.unicode_minus"] = False

XR_constants = {2: {"A2":1.880, "D3":0,     "D4":3.267},
                3:  {"A2":1.023, "D3":0,     "D4":2.574},
                4:  {"A2":0.729, "D3":0,     "D4":2.282},
                5:  {"A2":0.577, "D3":0,     "D4":2.114},
                6:  {"A2":0.483, "D3":0,     "D4":2.004},
                7:  {"A2":0.419, "D3":0.076, "D4":1.924},
                8:  {"A2":0.373, "D3":0.136, "D4":1.864},
                9:  {"A2":0.337, "D3":0.184, "D4":1.816},
                10: {"A2":0.308, "D3":0.223, "D4":1.777},
                11: {"A2":0.285, "D3":0.256, "D4":1.744},
                12: {"A2":0.266, "D3":0.283, "D4":1.717},
                13: {"A2":0.249, "D3":0.307, "D4":1.693},
                14: {"A2":0.235, "D3":0.328, "D4":1.672},
                15: {"A2":0.223, "D3":0.347, "D4":1.653},
                16: {"A2":0.212, "D3":0.363, "D4":1.637},
                17: {"A2":0.203, "D3":0.378, "D4":1.622},
                18: {"A2":0.194, "D3":0.391, "D4":1.608},
                19: {"A2":0.187, "D3":0.403, "D4":1.597},
                20: {"A2":0.180, "D3":0.415, "D4":1.585},
                21: {"A2":0.173, "D3":0.425, "D4":1.575},
                22: {"A2":0.167, "D3":0.434, "D4":1.566},
                23: {"A2":0.162, "D3":0.443, "D4":1.557},
                24: {"A2":0.157, "D3":0.451, "D4":1.548},
                25: {"A2":0.153, "D3":0.459, "D4":1.541}}
XS_constants = {2: {"A3":2.659, "B3":0,     "B4":3.267},
                3: {"A3":1.954, "B3":0,     "B4":2.568},
                4: {"A3":1.628, "B3":0,     "B4":2.266},
                5: {"A3":1.427, "B3":0,     "B4":2.089},
                6: {"A3":1.287, "B3":0.030, "B4":1.97},
                7: {"A3":1.182, "B3":0.118, "B4":1.882},
                8: {"A3":1.099, "B3":0.185, "B4":1.815},
                9: {"A3":1.032, "B3":0.239, "B4":1.761},
                10:{"A3":0.975, "B3":0.284, "B4":1.716},
                11:{"A3":0.927, "B3":0.321, "B4":1.679},
                12:{"A3":0.886, "B3":0.354, "B4":1.646},
                13:{"A3":0.850, "B3":0.382, "B4":1.618},
                14:{"A3":0.817, "B3":0.406, "B4":1.594},
                15:{"A3":0.789, "B3":0.428, "B4":1.572},
                16:{"A3":0.763, "B3":0.448, "B4":1.552},
                17:{"A3":0.739, "B3":0.466, "B4":1.534},
                18:{"A3":0.718, "B3":0.482, "B4":1.518},
                19:{"A3":0.698, "B3":0.497, "B4":1.503},
                20:{"A3":0.680, "B3":0.510, "B4":1.49},
                21:{"A3":0.663, "B3":0.523, "B4":1.477},
                22:{"A3":0.647, "B3":0.534, "B4":1.466},
                23:{"A3":0.633, "B3":0.545, "B4":1.455},
                24:{"A3":0.619, "B3":0.555, "B4":1.445},
                25:{"A3":0.606, "B3":0.565, "B4":1.435}}
IMR_constants = {"d2": 1.128, "D3": 0.0, "D4": 3.267}

def load_metric_data(xlsx_path, target_sheet):

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if target_sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Worksheet not found: {target_sheet}")

    ws = wb[target_sheet]
    rows = []
    for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
        vals = [np.nan if v is None else float(v) for v in row]
        if np.all(np.isnan(vals)):
            continue
        rows.append(vals)
    wb.close()

    if not rows:
        raise ValueError("No valid data in worksheet")
    return np.asarray(rows, dtype=float)

def get_XS_constant(n):

    table_n = np.array(list(XS_constants.keys()))

    # 超過最大值使用最大表值
    if n >= max(table_n):
        return XS_constants[max(table_n)]

    # 找最接近值
    idx = np.argmin(abs(table_n-n))

    return XS_constants[table_n[idx]]

def build_control_limit(data,chart = 'Xbar-R'):

    data=np.array(data)
    n=data.shape[1]
    xbar=np.mean(data,axis=1)
    xbarbar=np.mean(xbar)

    if chart == 'Xbar-R':
        R=np.ptp(data,axis=1)
        Rbar=np.mean(R)
        c=XR_constants[n]
        limits={"type":"Xbar-R",
                "UCL_X":xbarbar+c["A2"]*Rbar,
                "CL_X" :xbarbar,
                "LCL_X":xbarbar-c["A2"]*Rbar,
                "UCL_R":c["D4"]*Rbar,
                "CL_R" :Rbar,
                "LCL_R":c["D3"]*Rbar}

    elif chart == 'Xbar-S':
        S=np.std(data,axis=1,ddof=1)
        Sbar=np.mean(S)
        c=get_XS_constant(n)
        limits={"type" :"Xbar-S",
                "UCL_X":xbarbar+c["A3"]*Sbar,
                "CL_X" :xbarbar,
                "LCL_X":xbarbar-c["A3"]*Sbar,
                "UCL_S":c["B4"]*Sbar,
                "CL_S" :Sbar,
                "LCL_S":c["B3"]*Sbar}
    elif chart == 'I-MR':
        I = xbar
        MR = np.abs(np.diff(I))
        MRbar = np.mean(MR) if len(MR) > 0 else 0.0
        c = IMR_constants
        sigma = MRbar / c["d2"] if c["d2"] > 0 else 0.0
        limits={"type":"I-MR",
                "UCL_I":np.mean(I)+3*sigma,
                "CL_I" :np.mean(I),
                "LCL_I":np.mean(I)-3*sigma,
                "UCL_MR":c["D4"]*MRbar,
                "CL_MR" :MRbar,
                "LCL_MR":c["D3"]*MRbar}
    else:
        raise ValueError(f"Unsupported chart type: {chart}")
    return limits

def detect_control_rule_violations(values, cl, sigma):
    """
    Detect X-bar control chart rule violations.
    Returns:
        {
            "indices": {rule_no: [idx0, idx1, ...]},
            "messages": ["Rule 1 ...", ...]
        }
    """
    x = np.asarray(values, dtype=float)
    n = len(x)

    violations = {k: set() for k in range(1, 8)}

    # Rule 1 (單點超出 3σ 管制界限)
    idx = np.where((x > cl + 3*sigma) | (x < cl - 3*sigma))[0]
    violations[1].update(idx.tolist())

    # Rule 2 (連續 9 點在中心線同一側)
    for s in range(max(0, n - 9 + 1)):
        w = x[s:s+9]
        if np.all(w > cl) or np.all(w < cl):
            violations[2].update(range(s, s+9))

    # Rule 3 (連續 6 點遞增或遞減)
    for s in range(max(0, n - 6 + 1)):
        w = x[s:s+6]
        d = np.diff(w)
        if np.all(d > 0) or np.all(d < 0):
            violations[3].update(range(s, s+6))

    # Rule 4 (連續 14 點呈現交替上升下降)
    for s in range(max(0, n - 14 + 1)):
        w = x[s:s+14]
        d = np.diff(w)
        if np.all(d != 0) and np.all(np.sign(d[1:]) == -np.sign(d[:-1])):
            violations[4].update(range(s, s+14))

    # Rule 5 (連續 2 點超出 2σ 管制界限)
    for s in range(max(0, n - 2 + 1)):
        w = x[s:s+2]
        if np.sum(np.abs(w - cl) > 2*sigma) >= 1:
            violations[5].update(range(s, s+2))

    # Rule 6 (連續 4 點超出 1σ 管制界限)
    for s in range(max(0, n - 4 + 1)):
        w = x[s:s+4]
        if np.sum(np.abs(w - cl) > 1*sigma) >= 3:
            violations[6].update(range(s, s+4))

    # Rule 7 (連續 15 點在 1σ 管制界限內)
    for s in range(max(0, n - 15 + 1)):
        w = x[s:s+15]
        if np.all(np.abs(w - cl) <= 1*sigma):
            violations[7].update(range(s, s+15))

    messages = []
    sorted_idx = {}
    for r in sorted(violations.keys()):
        ids = sorted(violations[r])
        sorted_idx[r] = ids
        if ids:
            # Show 1-based sample index
            ids_1based = [i + 1 for i in ids]
            messages.append(f"Rule {r} violated points: {ids_1based}")

    return {"indices": sorted_idx, "messages": messages}

def annotate_point_values(ax, values, x_positions=None, default_color='black',
                          highlight_indices=None, highlight_color='red',
                          dy=8, fontsize=12):
    y = np.asarray(values, dtype=float)
    if x_positions is None:
        x_positions = np.arange(len(y))
    highlight_set = set(highlight_indices or [])
    for idx, (xi, yi) in enumerate(zip(x_positions, y)):
        txt_color = highlight_color if idx in highlight_set else default_color
        ax.annotate(f"{yi:.2f}", (xi, yi), textcoords='offset points', xytext=(0, dy),
                    ha='center', fontsize=fontsize, color=txt_color)

def annotate_control_limits(ax, ucl, cl, lcl, fontsize=9):
    trans = ax.get_yaxis_transform()  # x 用軸比例、y 用資料座標
    ax.text(1.01, ucl, f"UCL: {ucl:.2f}", transform=trans, va='center', ha='left', fontsize=fontsize)
    ax.text(1.01, cl, f"CL: {cl:.2f}", transform=trans, va='center', ha='left', fontsize=fontsize)
    ax.text(1.01, lcl, f"LCL: {lcl:.2f}", transform=trans, va='center', ha='left', fontsize=fontsize)

def plot_control_chart(sheet_name, data_path, chart_type='Xbar-R'):
    data = load_metric_data(data_path, sheet_name)

    # Phase I
    limits = build_control_limit(data[:25],chart = chart_type)
    phase = data.shape[0]//25+1
    i = data.shape[0]%25 if data.shape[0] > 25 else 25
    xbar = np.mean(data[-i:], axis=1)
    if chart_type == 'Xbar-R':
        R = np.max(data[-i:], axis=1) - np.min(data[-i:], axis=1)
    if chart_type == 'Xbar-S':
        S = np.std(data[-i:], axis=1, ddof=1)
    if chart_type == 'I-MR':
        I = xbar
        MR = np.abs(np.diff(I))

    if limits['type'] == 'I-MR':
        main_values = I
        main_cl = limits['CL_I']
        sigma_main = max((limits['UCL_I'] - limits['CL_I']) / 3, 1e-12)
        main_chart_title = 'I'
    else:
        main_values = xbar
        main_cl = limits['CL_X']
        sigma_main = max((limits['UCL_X'] - limits['CL_X']) / 3, 1e-12)
        main_chart_title = 'X-bar'

    rule_result = detect_control_rule_violations(main_values, main_cl, sigma_main)
    rule_indices = rule_result["indices"]
    sample_x = np.arange(1, len(main_values) + 1)
    mr_x = sample_x[1:]

    all_violation_idx = sorted(set().union(*[set(v) for v in rule_indices.values()]))

    # R/S 圖規則判定
    if limits['type'] == 'Xbar-R':
        sub_values = R
        sub_cl = limits['CL_R']
        sub_sigma = max((limits['UCL_R'] - limits['CL_R']) / 3, 1e-12)
        sub_chart_name = 'R'
    elif limits['type'] == 'Xbar-S':
        sub_values = S
        sub_cl = limits['CL_S']
        sub_sigma = max((limits['UCL_S'] - limits['CL_S']) / 3, 1e-12)
        sub_chart_name = 'S'
    else:
        sub_values = MR
        sub_cl = limits['CL_MR']
        sub_sigma = max((limits['UCL_MR'] - limits['CL_MR']) / 3, 1e-12)
        sub_chart_name = 'MR'

    sub_rule_result = detect_control_rule_violations(sub_values, sub_cl, sub_sigma)
    sub_rule_indices = sub_rule_result["indices"]
    sub_all_violation_idx = sorted(set().union(*[set(v) for v in sub_rule_indices.values()]))

    fig, ax = plt.subplots(2, 1, figsize=(12, 8))
    # Main Chart (X-bar or I)
    main_line, = ax[0].plot(sample_x, main_values, marker='o')
    if limits['type'] == 'I-MR':
        ax[0].axhline(limits['UCL_I'], linestyle='--', c = 'r')
        ax[0].axhline(limits['CL_I'], c = 'g')
        ax[0].axhline(limits['LCL_I'], linestyle='--', c = 'r')
    else:
        ax[0].axhline(limits['UCL_X'], linestyle='--', c = 'r')
        ax[0].axhline(limits['CL_X'], c = 'g')
        ax[0].axhline(limits['LCL_X'], linestyle='--', c = 'r')
    #ax[0].axhline(limits['CL_X'] + 1*sigma_x, linestyle=':', color='gray', alpha=0.7)
    #ax[0].axhline(limits['CL_X'] - 1*sigma_x, linestyle=':', color='gray', alpha=0.7)
    #ax[0].axhline(limits['CL_X'] + 2*sigma_x, linestyle='-.', color='gray', alpha=0.7)
    #ax[0].axhline(limits['CL_X'] - 2*sigma_x, linestyle='-.', color='gray', alpha=0.7)

    if all_violation_idx:
        ax[0].scatter(sample_x[all_violation_idx], main_values[all_violation_idx], color='red', s=70, zorder=5, label='Rule Warning')

    annotate_point_values(
        ax[0], main_values, x_positions=sample_x,
        default_color=main_line.get_color(),
        highlight_indices=all_violation_idx,
        highlight_color='red',
        dy=8, fontsize=8
    )

    ax[0].set_title(f'{main_chart_title} Control Chart - {sheet_name} (phase{phase})')
    ax[0].set_ylabel('Mean')
    ax[0].set_xlim(0.5, len(sample_x) + 0.5)
    ax[0].xaxis.set_major_locator(MaxNLocator(integer=True))
    if limits['type'] == 'I-MR':
        annotate_control_limits(ax[0], limits['UCL_I'], limits['CL_I'], limits['LCL_I'])
    else:
        annotate_control_limits(ax[0], limits['UCL_X'], limits['CL_X'], limits['LCL_X'])

    if rule_result["messages"]:
        xbar_log_lines = ["Control chart rule warnings:"] + rule_result["messages"]
    else:
        xbar_log_lines = ["Control chart rule warnings: none"]

    if sub_rule_result["messages"]:
        sub_log_lines = [f"{sub_chart_name}-chart rule warnings:"] + sub_rule_result["messages"]
    else:
        sub_log_lines = [f"{sub_chart_name}-chart rule warnings: none"]

    log_path = Path(data_path).with_name("control_chart_rule_warnings.txt")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] sheet={sheet_name}, phase={phase}\n")
        for line in xbar_log_lines:
            f.write(line + "\n")
        for line in sub_log_lines:
            f.write(line + "\n")
        f.write("-" * 80 + "\n")

    # R Chart
    if limits['type'] == 'Xbar-R':
        r_line, = ax[1].plot(sample_x, R, marker='o')
        ax[1].axhline(limits['UCL_R'], linestyle='--', c = 'r')
        ax[1].axhline(limits['CL_R'], c = 'g')
        ax[1].axhline(limits['LCL_R'], linestyle='--', c = 'r')

        if sub_all_violation_idx:
            ax[1].scatter(sample_x[sub_all_violation_idx], R[sub_all_violation_idx], color='red', s=70, zorder=5)

        ax[1].set_title(f'R Control Chart - {sheet_name}')
        ax[1].set_ylabel('Range')
        ax[1].set_xlabel('Sample')
        ax[1].set_xlim(0.5, len(sample_x) + 0.5)
        ax[1].xaxis.set_major_locator(MaxNLocator(integer=True))
        annotate_control_limits(ax[1], limits['UCL_R'], limits['CL_R'], limits['LCL_R'])
        annotate_point_values(
            ax[1], R, x_positions=sample_x,
            default_color=r_line.get_color(),
            highlight_indices=sub_all_violation_idx,
            highlight_color='red',
            dy=8, fontsize=8
        )
    # S Chart
    elif limits['type'] == 'Xbar-S':
        s_line, = ax[1].plot(sample_x, S, marker='o')
        ax[1].axhline(limits['UCL_S'], linestyle='--', c = 'r')
        ax[1].axhline(limits['CL_S'], c = 'g')
        ax[1].axhline(limits['LCL_S'], linestyle='--', c = 'r')

        if sub_all_violation_idx:
            ax[1].scatter(sample_x[sub_all_violation_idx], S[sub_all_violation_idx], color='red', s=70, zorder=5)

        ax[1].set_title(f'S Control Chart - {sheet_name}')
        ax[1].set_ylabel('Standard Deviation')
        ax[1].set_xlabel('Sample')
        ax[1].set_xlim(0.5, len(sample_x) + 0.5)
        ax[1].xaxis.set_major_locator(MaxNLocator(integer=True))
        annotate_control_limits(ax[1], limits['UCL_S'], limits['CL_S'], limits['LCL_S'])
        annotate_point_values(
            ax[1], S, x_positions=sample_x,
            default_color=s_line.get_color(),
            highlight_indices=sub_all_violation_idx,
            highlight_color='red',
            dy=8, fontsize=8
        )
    # MR Chart
    else: # limits['type'] == 'I-MR'
        mr_line, = ax[1].plot(mr_x, MR, marker='o')
        ax[1].axhline(limits['UCL_MR'], linestyle='--', c = 'r')
        ax[1].axhline(limits['CL_MR'], c = 'g')
        ax[1].axhline(limits['LCL_MR'], linestyle='--', c = 'r')

        if sub_all_violation_idx:
            ax[1].scatter(mr_x[sub_all_violation_idx], MR[sub_all_violation_idx], color='red', s=70, zorder=5)

        ax[1].set_title(f'MR Control Chart - {sheet_name}')
        ax[1].set_ylabel('Moving Range')
        ax[1].set_xlabel('Sample')
        ax[1].set_xlim(0.5, len(sample_x) + 0.5)
        ax[1].xaxis.set_major_locator(MaxNLocator(integer=True))
        annotate_control_limits(ax[1], limits['UCL_MR'], limits['CL_MR'], limits['LCL_MR'])
        annotate_point_values(
            ax[1], MR, x_positions=mr_x,
            default_color=mr_line.get_color(),
            highlight_indices=sub_all_violation_idx,
            highlight_color='red',
            dy=8, fontsize=8
        )

    output_dir = Path(data_path).parent
    safe_sheet_name = "".join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in sheet_name)
    fig_path = output_dir / f"control_chart_{safe_sheet_name}.png"
    draw_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    fig.text(0.995, 0.005, f"Drawn at: {draw_time_str}", ha='right', va='bottom', fontsize=8, color='gray')
    plt.tight_layout()
    fig.savefig(fig_path, dpi=300, bbox_inches='tight')
    plt.close(fig)

measure_folder = Path(__file__).resolve().parent
data_path = measure_folder / "control_chat_table.xlsx"
plot_control_chart('Short LB - Peak Loss', data_path, chart_type='Xbar-R')
plot_control_chart('Short LB - Peak Wavelength', data_path, chart_type='Xbar-R')
plot_control_chart('Short LB - Bandwidth', data_path, chart_type='Xbar-R')
plot_control_chart('Short LB - Loss at 1300', data_path, chart_type='Xbar-R')
plot_control_chart('Short LB - Loss at 1308', data_path, chart_type='Xbar-R')
plot_control_chart('Short LB - Loss at 1320', data_path, chart_type='Xbar-R')

plot_control_chart('G1 - Peak Loss', data_path, chart_type='I-MR')
plot_control_chart('G1 - Peak Wavelength', data_path, chart_type='I-MR')
plot_control_chart('R15 - Average Loss', data_path, chart_type='I-MR')
plot_control_chart('R16 - Average Loss', data_path, chart_type='I-MR')
